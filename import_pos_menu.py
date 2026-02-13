import argparse
import json
import re
import os
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except Exception as e:
    load_workbook = None

from odoo_jsonrpc import OdooClient, OdooRPCError


def load_config(path: str = "config.json") -> dict:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def guess_columns(headers: List[str]) -> Dict[str, int]:
    lowered = [h.strip().lower() for h in headers]

    def find(candidates: List[str]) -> Optional[int]:
        for cand in candidates:
            c = cand.strip().lower()
            for i, h in enumerate(lowered):
                if h == c or c in h:
                    return i
        return None

    name_idx = find(["name", "品名", "商品", "品項", "menu", "項目", "名稱", "主商品名稱"])
    price_idx = find(["price", "售價", "單價", "價格", "價", "list_price", "主商品價格"])
    category_idx = find(["category", "類別", "分類", "pos類別", "pos category", "主商品類別"])
    sku_idx = find(["sku", "code", "條碼", "barcode", "貨號", "型號", "主商品代碼", "主商品料號", "主商品編號"])
    combo_idx = find(["套用加購選單", "加購題型選單", "加購組合", "題型選項組合編號", "選項組合編號"])

    mapping: Dict[str, int] = {}
    if name_idx is not None: mapping["name"] = name_idx
    if price_idx is not None: mapping["price"] = price_idx
    if category_idx is not None: mapping["category"] = category_idx
    if sku_idx is not None: mapping["sku"] = sku_idx
    if combo_idx is not None: mapping["combo_id"] = combo_idx
    return mapping


def read_excel(path: str, sheet: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，請先安裝：pip install openpyxl")
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    mapping = guess_columns(headers)

    def get_val(r: List[Any], idx: Optional[int]) -> Any:
        if idx is None: return None
        return r[idx] if idx < len(r) else None

    items: List[Dict[str, Any]] = []
    for r in rows[1:]:
        name = get_val(r, mapping.get("name"))
        price = get_val(r, mapping.get("price"))
        category = get_val(r, mapping.get("category"))
        sku = get_val(r, mapping.get("sku"))
        combo_id = get_val(r, mapping.get("combo_id"))
        if name is None or (isinstance(name, str) and not name.strip()): continue
        p: Optional[float] = None
        if price is not None:
            try:
                if isinstance(price, str): price = re.sub(r"[^0-9.\-]", "", price)
                p = float(price)
            except Exception: p = None
        items.append({
            "name": str(name).strip(),
            "price": p,
            "category": str(category).strip() if isinstance(category, str) else category,
            "sku": str(sku).strip() if isinstance(sku, str) else sku,
            "combo_id": str(combo_id).strip() if isinstance(combo_id, str) else combo_id,
        })
    return items, headers


def _attribute_type_from_group(group_name: str) -> str:
    g = (group_name or "").strip().lower()
    if any(k in g for k in ["甜", "sweet"]): return "sweetness"
    if any(k in g for k in ["溫度", "冰", "熱", "temp", "temperature"]): return "temperature"
    if any(k in g for k in ["尺寸", "大小", "size"]): return "size"
    return "other"


def read_combo_options(path: str) -> Dict[str, List[Dict[str, Any]]]:
    if load_workbook is None: raise RuntimeError("缺少 openpyxl")
    wb = load_workbook(path, data_only=True)
    combos: Dict[str, str] = {}
    if "加購題型選項組合" in wb.sheetnames:
        ws = wb["加購題型選項組合"]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[1:]:
            cid = r[0]
            if cid: combos[str(cid)] = str(r[1] or "")
    mapping: Dict[str, List[Dict[str, Any]]] = {}
    if "加購選項項目" in wb.sheetnames:
        ws2 = wb["加購選項項目"]
        rows2 = list(ws2.iter_rows(values_only=True))
        for r in rows2[1:]:
            cid = r[0]
            group = r[1]
            opt_name = r[4] if len(r) > 4 else r[3]
            price_raw = r[5] if len(r) > 5 else None
            price: float = 0.0
            if price_raw is not None:
                try: price = float(str(price_raw).replace(",", ""))
                except Exception: price = 0.0
            if cid and opt_name:
                k = str(cid)
                mapping.setdefault(k, []).append({
                    "type": _attribute_type_from_group(str(group or "")),
                    "name": str(opt_name),
                    "price": price,
                })
    return mapping


def ensure_db(odoo: OdooClient, db: Optional[str]) -> str:
    if db: return db
    dbs = odoo.list_databases() or []
    if isinstance(dbs, dict) and dbs.get("databases"): dbs = dbs["databases"]
    if not dbs: raise RuntimeError("No databases found on server")
    return dbs[0]


def ensure_pos_category(odoo: OdooClient, name: str) -> int:
    cats = odoo.search_read("pos.category", [["name", "=", name]], ["id"], limit=1)
    if cats: return cats[0]["id"]
    return odoo.create("pos.category", {"name": name})


def create_or_update_product(odoo: OdooClient, name: str, list_price: Optional[float], category_id: Optional[int], update_existing: bool = True) -> Tuple[int, str]:
    existing = odoo.search_read("product.template", [["name", "=", name]], ["id", "list_price", "available_in_pos"], limit=1)
    pos_cat_field = "pos_category_id"

    def vals(base_price: Optional[float], cat_id: Optional[int]) -> Dict[str, Any]:
        v: Dict[str, Any] = {"name": name, "sale_ok": True, "available_in_pos": True, "type": "consu"}
        if base_price is not None: v["list_price"] = base_price
        if cat_id is not None: v[pos_cat_field] = cat_id
        return v

    if existing:
        pid = existing[0]["id"]
        action = "skipped"
        if update_existing:
            try:
                odoo.write("product.template", [pid], vals(list_price, category_id))
                action = "updated"
            except OdooRPCError:
                pos_cat_field = "pos_categ_id"
                odoo.write("product.template", [pid], vals(list_price, category_id))
                action = "updated"
        return pid, action

    try:
        new_id = odoo.create("product.template", vals(list_price, category_id))
        return new_id, "created"
    except OdooRPCError:
        pos_cat_field = "pos_categ_id"
        new_id = odoo.create("product.template", vals(list_price, category_id))
        return new_id, "created"


def main() -> None:
    cfg = load_config()
    odoo_cfg = cfg.get("odoo", {})
    ap = argparse.ArgumentParser(description="Import POS menu from Excel into Odoo")
    ap.add_argument("--source", required=True, help="Excel file path (.xlsx)")
    ap.add_argument("--sheet", help="Sheet name (default: active)")
    ap.add_argument("--url", default=odoo_cfg.get("url", "http://127.0.0.1:8069"), help="Odoo base URL")
    ap.add_argument("--login", default=odoo_cfg.get("login", "admin"), help="Odoo login")
    ap.add_argument("--password", default=odoo_cfg.get("password", "admin"), help="Odoo password")
    ap.add_argument("--db", default=odoo_cfg.get("db"), help="Database name (optional)")
    ap.add_argument("--apply", action="store_true", help="Apply changes to Odoo (default: dry-run)")
    ap.add_argument("--update-existing", action="store_true", help="Update existing products")
    ap.add_argument("--skip-ambiguous", action="store_true", help="Skip items with uncertain fields")
    ap.add_argument("--no-config", action="store_true", help="Do not create beverage config")
    ap.add_argument("--only-combo", action="store_true", help="Only include items with combo")
    ap.add_argument("--dump", help="Write prepared payload JSON")
    args = ap.parse_args()

    items, headers = read_excel(args.source, args.sheet)
    combo_map = read_combo_options(args.source)
    if not items:
        print("[error] No rows parsed.")
        return

    if not args.apply:
        print("[dry-run] Use --apply to write to Odoo.")
        return

    client = OdooClient(args.url)
    db = ensure_db(client, args.db)
    client.authenticate(db, args.login, args.password)
    cat_cache: Dict[str, int] = {}

    for it in items:
        name = str(it.get("name") or "").strip()
        if not name: continue
        price = it.get("price")
        cat_name = str(it.get("category") or "").strip()
        combo_id = str(it.get("combo_id") or "").strip()

        cat_id: Optional[int] = None
        if cat_name:
            cat_id = cat_cache.get(cat_name)
            if cat_id is None:
                cat_id = ensure_pos_category(client, cat_name)
                cat_cache[cat_name] = cat_id

        sizes = [ln for ln in (combo_map.get(combo_id) or []) if ln.get("type") == "size"]
        m_extra = next((ln.get("price") or 0.0 for ln in sizes if str(ln.get("name")).strip().upper() == "M"), 0.0)
        price_m = (price or 0.0) + m_extra

        pid, action = create_or_update_product(client, name=name, list_price=price_m, category_id=cat_id, update_existing=args.update_existing)
        print(f"[{action}] product.template(id={pid}) name='{name}'")

        if not args.no_config and combo_id:
            # 檢查是否已存在配置，避免重複建立
            existing_cfg = client.search_read("pos.beverage.config", [["product_tmpl_id", "=", pid]], ["id"], limit=1)
            lines = combo_map.get(combo_id) or []
            cmd = [(5, 0, 0)]
            m_extra_cfg = next((ln.get("price") or 0.0 for ln in lines if ln.get("type") == "size" and str(ln.get("name")).strip().upper() == "M"), 0.0)
            for ln in lines:
                is_size = ln["type"] == "size"
                is_m = is_size and str(ln.get("name")).strip().upper() == "M"
                adj_price = (ln.get("price") or 0.0) - (m_extra_cfg if is_size else 0.0)
                cmd.append((0, 0, {"attribute_type": ln["type"], "name": ln["name"], "selected": bool(is_m), "price": adj_price}))
            vals = {"name": f"{name}設定", "product_tmpl_id": pid, "show_popup": True, "line_ids": cmd}
            if cat_id: vals["pos_category_id"] = cat_id

            if existing_cfg:
                client.write("pos.beverage.config", [existing_cfg[0]["id"]], vals)
                print(f"  [updated] beverage config for {name}")
            else:
                client.create("pos.beverage.config", vals)
                print(f"  [created] beverage config for {name}")

if __name__ == "__main__":
    main()
